import xlsxwriter
from openpyxl import load_workbook
import json



data = [
    {
        'Name' : "Khan Zikra",
        'Phone' : "9347296535",
        'Email' : "zikrakhan45@gmail.com",
        'Address' : "Khairani Road sakinaka Mumbai",
        'Country' : "India"
    },

    {
        'Name' : "Khan Shama",
        'Phone' : "9347296387",
        'Email' : "shamakhan32@gmail.com",
        'Address' : "shangarh nagar chandivali Mumbai",
        'Country' : "India"
    },
    {
        'Name' : "khan Aalma",
        'Phone' : "434729784",
        'Email' : "aalma33@gmail.com",
        'Address' : "Khairani Road chandivali",
        'Country' : "India"

    },
    {
        'Name' : "Khan Bushra",
        'Phone' : "5678436953",
        'Email' : "bushra1222@gmail.com",
        'Address' : "Khairani Road chandivali Mumbai",
        'Country' : "America"
    },
    {
        'Name' : "Manihar Aafreen",
        'Phone' : "56784555",
        'Email' : "bushra1222@gmail.com",
        'Address' : "Khairani Road chandivali Mumbai",
        'Country' : "Tilak Nagar"
    }
] 

workbook = xlsxwriter.Workbook("AllAboutPythonExcel.xlsx")
worksheet = workbook.add_worksheet("firstSheet")

worksheet.write(0,0, "#")    # 1st  Argument Row No , 2nd Argument Column No, 3rd Argument Text You want to enter 
worksheet.write(0,1,"Name") 
worksheet.write(0,2,"Phone")
worksheet.write(0,3,"Email")
worksheet.write(0,4,"Address")
worksheet.write(0,5,"Country")

for index , entry in enumerate(data):
    worksheet.write(index+1 , 0, str(index))
    worksheet.write(index+1, 1, entry["Name"])
    worksheet.write(index+1, 2, entry["Phone"])
    worksheet.write(index+1, 3, entry["Email"])
    worksheet.write(index+1, 4, entry["Address"])
    worksheet.write(index+1, 5, entry["Country"])


workbook.close()



 # to read excelfile
wb = load_workbook(filename='AllAboutPythonExcel.xlsx')
sheet = wb['firstSheet']
print(sheet)

# 1 Retrieving cell value
wb = load_workbook(filename='AllAboutPythonExcel.xlsx')
sheet = wb['firstSheet']
# print(sheet['B3'].value)       # To get a specifid cell data 


nc = sheet['B2']                 # get a specifid cell data and store it in a variable 
print(nc.coordinate)

sheet = wb['firstSheet']
print(sheet.cell(row =2, column=2).value)


# 2 Retriveing Multiple Values
sheet = wb['firstSheet']
# print(sheet['A:C'])               # Get multiple data in coloumn     # Using both methode data showing in SHEETNAME for 
# print(sheet[1:3])                   # Get multiple data in row

for row in sheet["A1:F6"]:            # This method give values  
    print ([x.value for x in row])


# 3  Reading all the data in Excel

wb = load_workbook(filename='AllAboutPythonExcel.xlsx')
sheet = wb['firstSheet']
for row in sheet:
    print([data.value for data in row])





# 4 converting data into python structures
wb = load_workbook(filename='AllAboutPythonExcel.xlsx')
sheet = wb['firstSheet']

books = {}

for row in sheet.iter_rows(min_row=2,min_col=1, values_only= True):
    book_id  = row[0]
    book = {
        "Name":row[1],
        "Phone":row[2],
        "Email":row[3],
        "Address":row[4],
        "Country":row[5],

    }
    
    books[book_id] = book

print(json.dumps(books, indent= 4))        # indent means in wich quantity you want to see data


















